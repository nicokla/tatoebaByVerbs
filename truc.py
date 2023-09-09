

languageToCodes={
	'arabic':['ar','ar-EG','ar-SA','ar-LB','ar-MA','ar-SY','ar-IQ','ar-JO'],
	'chinese':['zh','zh-CN','zh-TW','zh-CHS','zh-Hans','zh-HK','zh-MO','zh-Hant','zh-CHT','zh-SG'],
	'tagalog':['tl','tl-PH'],
	'french':['fr','fr-FR','fr-CA','fr-BE','fr-CH'],
	'greek':['el','el-GR'],
	'hebrew':['he','he-IL','iw','iw-IL'],
	'hindi':['hi','hi-IN'],
	'italian':['it','it-IT'],
	'japanese':['ja','ja-JP'],
	'korean':['ko','ko-KR'],
	'persian':['fa','fa-IR'],
	'portuguese':['pt','pt-BR','pt-PT'],
	'russian':['ru','ru-RU'],
	'spanish':['es','es-AR','es-ES','es-MX'],
	'thai':['th','th-TH'],
	'turkish':['tr','tr-TR'],
	'vietnamese':['vi','vi-VN'],
	'english':['en','en-CA','en-US','en-GB','en-AU']
}


def getCode(langue):
	if(langue in languageToCodes):
		return languageToCodes[langue][0]
	else:
		return langue[:2]

codeToLanguage={}
for key in languageToCodes:
	for stuff in languageToCodes[key]:
		codeToLanguage[stuff]=key

def getLanguage(code):
	codeSmall=code[:2]
	return codeToLanguage[codeSmall]


# -------------------------
# 1) arabic

ar2la=[[u'لإ',u'il'],[u'لأ',u'al'],[u'إ',u'i'],[u'أ',u'a'],[u'ا',u'a'],[u'ر',u'r'],[u'ز',u'z'],[u'ش',u'š'],[u'س',u's'],[u'ص',u'ṡ'],[u'ض',u'ḋ'],[u'ط',u'ṫ'],[u'ظ',u'ż'],[u'ع',u'3'],[u'غ',u'ġ'],[u'ح',u'7'],[u'خ',u'ḣ'],[u'ج',u'j'],[u'ت',u't'],[u'ث',u'ṭ'],[u'د',u'd'],[u'ذ',u'ḍ'],[u' و',u' w'],[u'و',u'w'],[u'ي',u'y'],[u'ق',u'q'],[u'ف',u'f'],[u'ة',u'a'],[u'ه',u'h'],[u'ك',u'k'],[u'ل',u'l'],[u'م',u'm'],[u'ب',u'b'],[u'ن',u'n'],[u'ى',u'a'],[u'ُ',u'u'],[u'ِ',u'i'],[u'َ',u'a'],[u'؟',u'?'],[u'ْ',u''],[u'ّ',u'*'],[u'،',u','],[u'ً',u''],[u'ؤ',u'u'],[u'آ',u'aa'],[u'آ',u'aa'],[u'ء',u"'"],[u'ئ',u"'"],[u'١',u"1"],[u'٢',u"2"],[u'٣',u"3"],[u'٤',u"4"],[u'٥',u"5"],[u'٦',u"6"],[u'٧',u"7"],[u'٨',u"8"],[u'٩',u"9"],[u'٠',u"0"]]
# ,[u'٠',u"0"]

def transliterateArabic(string):
	for k in ar2la:
		string=string.replace(k[0], k[1])
	return string


# -----------------
# 2) japonais


import cutlet
katsu = cutlet.Cutlet()
katsu.use_foreign_spelling = False


# せっかく 神奈川まで来たので、鎌倉で 一泊することにしました。
# phrase1='votre {m} {f} vos {m-p} {f-p} [formal]'
# phrase2='[generic] (discourteous if used for a superior) 貴方達'
# transliterateJapaneseOld
def transliterateJapanese(phrase):
	currentJapanese=''
	result=''
	try:
		for c in phrase:
			if (c >= 'A' and c <= 'z') or (c in ' [].!?,;}{()'):
				if(currentJapanese!=''):
					result += katsu.romaji(currentJapanese).lower()
					currentJapanese=''
					if c == ' ' and not (result[-1] in '.!?,;'):
						result+=','
				result += c
			else:
				currentJapanese+=c
		if(currentJapanese!=''):
			result += katsu.romaji(currentJapanese).lower()
			currentJapanese=''
		return result
	except Exception as e:
		return phrase


# -----------------
# 3) chinois



import MicroTokenizer
import pinyin


def transliterateChineseSub(blob):
	try:
		return pinyin.get(blob, delimiter="")
	except Exception as e:
		return blob

def isLatin(mychar):
  number=ord(mychar)
  return number < 0x80

from functools import reduce

def maFonction(acc, s):
	if(len(acc)==0):
		return s
	if(len(s)==1 and isLatin(s[0]) and isLatin(acc[-1])):
		return acc+s
	else:
		return acc+' '+s

def transliterateChinese(phrase):
	tokens = MicroTokenizer.cut(phrase)
	liste=list(map(transliterateChineseSub,tokens)) 
	result = reduce(maFonction, liste, '') #' '.join(liste)
	return result

# -------------
# 4) russe



ru2la = [[u"о",u'o'],[u"О",u'O'],[u'Щ',u'Shtsh'],[u'щ',u'shtsh'],[u'Х',u'Kh'],[u'х',u'kh'],[u'Ч',u'Tsh'],[u'ч',u'tsh'],[u'Ш',u'Sh'],[u'ш',u'sh'],[u'Ц',u'Ts'],[u'ц',u'ts'],[u'Ё',u'Io'],[u'ё',u'io'],[u'е',u'ie'],[u'ю',u'iu'],[u'Е',u'Ie'],[u'Ю',u'Iu'],[u'Я',u'Ia'],[u'я',u'ia'],[u'а',u'a'],[u'з',u'z'],[u'э',u'e'],[u'р',u'r'],[u'т',u't'],[u'у',u'u'],[u'и',u'i'],[u'п',u'p'],[u'с',u's'],[u'д',u'd'],[u'Ф',u'f'],[u'г',u'g'],[u'ж',u'j'],[u'к',u'k'],[u'л',u'l'],[u'м',u'm'],[u'в',u'v'],[u'б',u'b'],[u'н',u'n'],[u'й',u'y'],[u'А',u'A'],[u'З',u'Z'],[u'Э',u'E'],[u'Р',u'R'],[u'Т',u'T'],[u'У',u'U'],[u'И',u'I'],[u'П',u'P'],[u'С',u'S'],[u'Д',u'D'],[u'ф',u'F'],[u'Г',u'G'],[u'Ж',u'J'],[u'К',u'K'],[u'Л',u'L'],[u'М',u'M'],[u'В',u'V'],[u'Б',u'B'],[u'Н',u'N'],[u'Й',u'Y']]


la2ru = []
for l in ru2la:
	la2ru.append([l[1],l[0]])

def latinToRus(s):
	for k in la2ru:
		s = s.replace(k[0], k[1])
	return s

def transliterateRussian(s):
	for k in ru2la:
		s = s.replace(k[0], k[1])
	return s

# -------------
# 5) hindi


import indicate
# english_translated = transliterate.hindi2english("हिंदी")
# print(english_translated)

def transliterateHindi(text):
	try:
		return indicate.transliterate.hindi2english(text)
	except Exception as e:
		return text


# -------------
# 6) hebrew

# [u'א',u'@'], [u'א,',u','], [u'א.',u'.'], [u'א?',u'?'],[u'א!',u'!'], [u' א',u' '],[u',א',u','], [u'.א',u'.'],[u'?א',u'?'],[u'!א',u'!'], [u' א',u' A']
# [u' ע',u' '], [u',ע',u','], [u'.ע',u'.'], [u'?ע',u'?'], [u'!ע',u'!'], [u'עַ ',u'a '],[u'עַ,',u'a,'],[u'עַ.',u'a.'],[u'עַ!',u'a!'],[u'עַ?',u'a?'],[u'ע ',u""],[u'ע,',u","],[u'ע.',u"."],[u'ע?',u"?"],[u'ע!',u"!"],[u'ע',u"'"]
# [u'ה ',u' '],[u'ה,',u','],[u'ה.',u'.'],[u'ה?',u'?'], [u'ה!',u'!']
he2la2 = [[u'א',u'@'],[u"ג'",u'j'], [u' נְ',u' ne'], [u' מְ',u' me'], [u' לְ',u' le'], [u' וְ',u' ve'],[u' יְ',u' ye'],[u' רְ',u' re'],[u' וֵּ',u've'],[u' תְּ',u'te'],[u' נְּ',u'ne'],[u' מְּ',u'me'],[u'ז',u'z'],[u'כָל ',u'ḣol '],[u'כָּל ',u'kol '],[u'ר',u'r'],[u'ת',u't'],[u'יֽ',u''], [u'יי',u'y'],[u'ִי',u'i'], [u'יִ',u'i'],[u'י',u'y'],[u'ִ',u'i'],[u'ק',u'k'],[u'ד',u'd'],[u'ג',u'g'],[u'ל',u'l'],[u'ט',u't'],[u'ס',u's'],[u'מ',u'm'],[u'ם',u'm'],[u'נ',u'n'],[u'ן',u'n'],[u'צ',u'ts'],[u'ץ',u'ts'],[u'וו',u'v'],[u'ֹו',u'o'], [u'וֹ',u'o'],[u'וּ',u'u'],[u'וֽ',u''], [u'ון',u'on'],[u'ו',u'w'],[u'שׁ',u'š'],[u'שׂ',u's'],[u'ש',u'š'],[u'פּ',u'p'],[u'פ',u'p'],[u'ף',u'f'],[u'כּ',u'k'],[u'כ',u'ḣ'],[u'ך',u'ḣ'],[u'בְּ',u'be'],[u'בּ',u'b'],[u'ב ',u'v '],[u'ב',u'b'], [u'ה',u'h'],[u'חַ,',u'aḣ,'],[u'חַ ',u'aḣ '],[u'חַ!',u'aḣ!'],[u'חַ?',u'aḣ?'],[u'חַ.',u'aḣ.'],[u'ח',u'ḣ'],[u'ע',u"3"],[u'ַ',u'a'],[u'ֱ',u'e'],[u'ֲ',u'a'],[u'ֳ',u'a'],[u'ָ',u'a'],[u'ֵ',u'e'],[u'ֶ',u'e'],[u'ֻ',u'u'],[u'ֹ',u'o'],[u'ְ',u''],[u'ּ',u''],[u'\\u05bd',u'']]


def transString(langue2lat, string):
	string=' '+string+' '
	for k in langue2lat:
		string=string.replace(k[0], k[1])
	return string.strip()

# def enleveNonLatin(mot):
# 	mot2=""
# 	for c in mot:
# 		if(ord(c)<128):
# 			mot2+=c
# 	return mot2

# def lastStep(phrase):
# 	phrase=phrase.lower() # can comment out if wanna see the vowels
# 	phrase=phrase.replace('|', ' ')
# 	return phrase

def transliterateHebrew(s):
	try:
		# return lastStep(enleveNonLatin(transString(he2la2, s)))
		return transString(he2la2, s)
	except Exception as e:
		return s


#  --------------

def transliterate(phrase, languageCode):
	language = getLanguage(languageCode)
	if language == 'japanese':
		return transliterateJapanese(phrase)
	elif language == 'arabic':
		return transliterateArabic(phrase)
	elif language == 'chinese':
		return transliterateChinese(phrase)
	elif language == 'russian':
		return transliterateRussian(phrase)
	elif language == 'hindi':
		return transliterateHindi(phrase)
	elif language == 'hebrew':
		return transliterateHebrew(phrase)
	else:
		return phrase


def transliterate2(word, langue):
	return transliterate(word, getCode(langue))

# ----------------


filesNames = dict()

filesNames['ar']="arabe/tatoeba ara_en.xlsx"
filesNames['ar-MA']="darija/tatoeba en_darija.xlsx"
filesNames['ja']="japonais/tatoeba fr_jap.xlsx"
filesNames['zh']="chinese/tatoeba fr_zh.xlsx"
filesNames['ru']="russian/tatoeba fr_rus.xlsx"
filesNames['hi']="hindi/tatoeba en_hindi.xlsx"
filesNames['he']="hebreu/tatoeba en_heb.xlsx"

myLanguageCodes=['ar','ar-MA','ja','zh','ru','hi','he']

import openpyxl
# from openpyxl.utils import coordinate_from_string, column_index_from_string
import os
os.chdir("/Users/nicolas/Desktop/interets/langues")

# languageCode=myLanguageCodes[0]
# myLanguageCodes
for languageCode in ['hi']: 
	fileName = filesNames[languageCode]
	wb1 = openpyxl.load_workbook(fileName)
	sheet1=wb1.worksheets[0]
	row_count=sheet1.max_row
	column_count=sheet1.max_column
	for b in range(1,row_count+1):
		print((b/row_count)*100)
		beforeTranscript = sheet1.cell(row=b,column=1).value
		afterTranscript = transliterate(beforeTranscript, languageCode)
		sheet1.cell(row=b,column=3).value = afterTranscript
	wb1.save(fileName)


# -----------------------

import openpyxl
import os
os.chdir("/Users/nicolas/Desktop/interets/langues")

# 1) absorb a xlsx
def getColonne(fileName, indexColonne):
	wb1 = openpyxl.load_workbook(fileName)
	sheet1=wb1.worksheets[0]
	colonne = []
	row_count=sheet1.max_row
	for b in range(1,row_count+1):
		value = sheet1.cell(row=b,column=indexColonne).value
		colonne.append(value)
	return colonne


# 2.0.1) getVerbes fr
import spacy
# from spacy.lang.fr.examples import sentences
# phrase = sentences[0]

nlp_fr = spacy.load('fr_core_news_sm')
nlp_en = spacy.load('en_core_web_sm')

phrase = 'I want to help you but I needed two small chicken and I thought about a thought.'

def getVerbes_en(phrase):
	doc = nlp_en(phrase) # doc.text == phrase
	l = set()
	for token in doc: 
		if token.pos_ == 'VERB':
			l.add(token.lemma_) # token.text, token.pos_
	return l

def getVerbes_fr(phrase):
	doc = nlp_fr(phrase) # doc.text == phrase
	l = set()
	for token in doc: 
		if token.pos_ == 'VERB':
			l.add(token.lemma_) # token.text, token.pos_
	return l

def getVerbes(phrase, language):
	if language == 'fr':
		return getVerbes_fr(phrase)
	else: # if language == 'en':
		return getVerbes_en(phrase)

# 2) detect all verbs

filesNames={
	'ka':'kabyle/tatoeba kabyle.xlsx',
	# 'ar':"arabe/tatoeba ara_en.xlsx",
	# 'ar-MA':"darija/tatoeba en_darija.xlsx",
	# 'ja':"japonais/tatoeba fr_jap.xlsx",
	# 'zh':"chinese/tatoeba fr_zh.xlsx",
	# 'ru':"russian/tatoeba fr_rus.xlsx",
	# 'hi':"hindi/tatoeba en_hindi.xlsx",
	# 'he':"hebreu/tatoeba en_heb.xlsx"
}

startLanguage={
	'ar':"en", 'ar-MA':"en", 'ja':"fr", 'zh':"fr",
	'ru':"fr",'hi':"en",'he':"en", 'ka':'fr'
}

outputNames={
	'ka':'kabyle/fr_ka_phrases par verbes.xlsx',
	'ar' : "arabe/en_ara_phrases par verbes.xlsx",
	'ar-MA' : "darija/en_darija_phrases par verbes.xlsx",
	# 'ja' : "japonais/tatoeba fr_jap.xlsx",
	'zh' : "chinese/fr_zh_phrases par verbes.xlsx",
	'ru' : "russian/fr_rus_phrases par verbes.xlsx",
	'hi' : "hindi/en_hin_phrases par verbes.xlsx",
	'he' : "hebreu/en_heb_phrases par verbes.xlsx"
}


def saveDico(dico, fileName):
	wb1 = openpyxl.Workbook()
	sheet1=wb1.worksheets[0]
	index = 1
	keys = dico.keys()
	sortedkeys = sorted(keys)
	for verbe in sortedkeys:
		for (phrase, phraseInconnue) in dico[verbe]:
			sheet1.cell(row=index,column=1).value = verbe
			sheet1.cell(row=index,column=2).value = phrase
			sheet1.cell(row=index,column=3).value = phraseInconnue
			index += 1
	wb1.save(fileName)

# nameFile = filesNames['ja']
toto = 0
for language, nameFile in filesNames.items():
	phrasesLangueConnue = getColonne(nameFile, 2)
	phrasesLangueInconnue = getColonne(nameFile, 3)
	dico = dict()
	length=len(phrasesLangueConnue)
	for index, phraseConnue in enumerate(phrasesLangueConnue):
		print(toto/6, index/length*100)
		phraseInconnue = phrasesLangueInconnue[index]
		verbes = getVerbes(phraseConnue, startLanguage[language])
		for verbe in verbes:
			if (verbe not in dico):
				dico[verbe] = [(phraseConnue, phraseInconnue)]
			else:
				dico[verbe].append((phraseConnue, phraseInconnue))
	saveDico(dico, outputNames[language])
	toto+=1

