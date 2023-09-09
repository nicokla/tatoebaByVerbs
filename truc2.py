
# fileName = '/Users/nicolas/Desktop/interets/langues/kabyle/fr_ka_phrases par verbes.xlsx'
# import openpyxl
# wb1 = openpyxl.load_workbook(fileName)
# sheet1=wb1.worksheets[0]
# row_count=sheet1.max_row

# liste = []
# newValue=''
# for row in range(1,row_count+1):
# 	oldValue = newValue
# 	newValue = sheet1.cell(row=row,column=1).value
# 	if newValue != oldValue:
# 		liste.append(newValue)

# import pickle
# def save_object(obj, filename):
#   with open(filename, 'wb') as output:  # Overwrites any existing file.
#     pickle.dump(obj, output, pickle.HIGHEST_PROTOCOL)

# save_object(liste, '/Users/nicolas/Desktop/interets/langues/kabyle/mesVerbes.pkl')

# -----------------------

import pickle

def get_object(filename):
  with open(filename, 'rb') as input:
    return pickle.load(input)

listeVerbesFr = get_object('/Users/nicolas/Desktop/interets/langues/kabyle/mesVerbes.pkl')

# --------------------------

# from algoliasearch.search_client import SearchClient
# client = SearchClient.create('IB3IUKX206', 'ff9605a51e2f1344a298c5a8ba1290f6')
# index = client.init_index('prod_amyag')

# monDico=dict()
# for verbe in listeVerbesFr:
# 	monDico[verbe] = []

# total = len(listeVerbesFr)
# for i, verbe in enumerate(listeVerbesFr):
# 	print(int(i/total*100))
# 	objects = index.search(verbe)
# 	hits = objects['hits']
# 	for hit in hits:
# 		monDico[verbe].append((hit['verb'], hit['senses']))

# import pickle
# def save_object(obj, filename):
#   with open(filename, 'wb') as output:  # Overwrites any existing file.
#     pickle.dump(obj, output, pickle.HIGHEST_PROTOCOL)

# save_object(monDico, '/Users/nicolas/Desktop/interets/langues/kabyle/mesVerbesKabyles.pkl')

# -----------------------------

listeVerbesKabyles = get_object('/Users/nicolas/Desktop/interets/langues/kabyle/mesVerbesKabyles.pkl')

#____________________________________________________

import openpyxl
wb1 = openpyxl.Workbook()
sheet1=wb1.worksheets[0]

ligne = 1
for i, verbeFr in enumerate(listeVerbesKabyles):
	verbesKabyles = listeVerbesKabyles[verbeFr]
	for verbeKa in verbesKabyles:
		leVerbe = verbeKa[0]
		leSens = verbeKa[1]
		sheet1.cell(row=ligne,column=1).value = verbeFr
		sheet1.cell(row=ligne,column=2).value = leVerbe
		sheet1.cell(row=ligne,column=3).value = leSens
		ligne += 1

fileName = '/Users/nicolas/Desktop/interets/langues/kabyle/verbesKabyles.xlsx'
wb1.save(fileName)

# ------------------------


import openpyxl
wb1 = openpyxl.Workbook()
sheet1=wb1.worksheets[0]


dico2=dict()
for i, verbeFr in enumerate(listeVerbesKabyles):
	verbesKabyles = listeVerbesKabyles[verbeFr]
	for verbeKa in verbesKabyles:
		leVerbe = verbeKa[0]
		leSens = verbeKa[1]
		dico2[leVerbe] = leSens

for i, leVerbe in enumerate(dico2):
	sheet1.cell(row=i+1,column=1).value = leVerbe
	sheet1.cell(row=i+1,column=2).value = dico2[leVerbe]


fileName = '/Users/nicolas/Desktop/interets/langues/kabyle/verbesKabylesFromKa.xlsx'
wb1.save(fileName)

